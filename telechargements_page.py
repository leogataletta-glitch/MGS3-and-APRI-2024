"""Onglet « Téléchargement des données ».

Six jeux à télécharger, produits à la demande à partir des mêmes fichiers que
le tableau de bord — jamais d'export figé à côté, qui finirait par diverger de
ce qui est affiché. Chaque classeur s'ouvre sur une feuille « Lisez-moi » qui
dit d'où viennent les chiffres et comment lire les colonnes.
"""

import io
import json
import os
import pickle

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import i18n
from i18n import T

APP_DIR = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(APP_DIR, "data")

SECTIONS = ["Anse à Drick", "Barbois", "Dumont", "Débouchette", "Mouline",
            "Quentin", "Beaulieu", "Blactote", "Dalmette", "Trichet"]
SOUS_POP = ["Total", "Homme", "Femme", "Cat A", "Cat B", "Cat C",
            "<25", "25-39", "40-59", "60+"]
PAYSAGES = ["Littoral", "Montagne"]
GROUPES = SOUS_POP + PAYSAGES + SECTIONS

DIM_ORDRE = [
    "I. PHYSICAL AND INFRASTRUCTURAL DIMENSION",
    "II. INSTITUTIONAL, TECHNOLOGICAL, AND GOVERNANCE  DIMENSION",
    "III.  ENVIRONMENTAL AND ECOLOGICAL DIMENSION",
    "IV. ECONOMIC, LIVELIHOODS, AND FOOD SECURITY DIMENSION",
    "V. SOCIAL AND COMMUNITY DIMENSION",
    "VI. HUMAN DIMENSION",
    "VII. CULTURAL, IDENTITY-BASED, AND PSYCHOLOGICAL DIMENSION",
]

POLICE = "Arial"
ENTETE_FOND = PatternFill("solid", fgColor="1A6BB0")
ENTETE_ENCRE = Font(name=POLICE, size=10, bold=True, color="FFFFFF")
NORMAL = Font(name=POLICE, size=10)
GRAS = Font(name=POLICE, size=10, bold=True)
TITRE = Font(name=POLICE, size=13, bold=True, color="101728")
FILET = Border(bottom=Side("thin", color="D8DEE8"))


def _chemin(nom):
    for c in (os.path.join(DATA, nom), os.path.join(APP_DIR, nom)):
        if os.path.exists(c):
            return c
    return None


def _lire_json(nom):
    c = _chemin(nom)
    if c is None:
        return None
    with open(c, encoding="utf-8") as f:
        return json.load(f)


def _entete(ws, ligne, valeurs, largeurs=None):
    for j, v in enumerate(valeurs, 1):
        c = ws.cell(row=ligne, column=j, value=v)
        c.font = ENTETE_ENCRE
        c.fill = ENTETE_FOND
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[ligne].height = 30
    if largeurs:
        for j, w in enumerate(largeurs, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(row=ligne + 1, column=1).coordinate


def _lisezmoi(wb, titre, lignes):
    """Feuille d'ouverture : ce que contient le classeur, et sa provenance."""
    ws = wb.create_sheet(T("x_lisezmoi")[:31], 0)
    ws.column_dimensions["A"].width = 118
    ws.cell(row=1, column=1, value=titre).font = TITRE
    r = 3
    for ligne in lignes:
        c = ws.cell(row=r, column=1, value=ligne)
        c.font = GRAS if ligne.endswith(":") else NORMAL
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(15, 15 * (len(ligne) // 100 + 1))
        r += 1
    ws.cell(row=r + 1, column=1, value=T("x_confid")).font = GRAS
    ws.cell(row=r + 1, column=1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r + 1].height = 45
    return ws


def _octets(wb):
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _nouveau():
    wb = Workbook()
    wb.remove(wb.active)
    return wb


# ----------------------------------------------------------------------
# 1 · résultats descriptifs
# ----------------------------------------------------------------------
@st.cache_data(show_spinner=False)
def _fichier_descriptif(lang):
    with open(_chemin("cache_national.pkl"), "rb") as f:
        cache = pickle.load(f)
    base_n = cache["base_n"]
    wb = _nouveau()
    _lisezmoi(wb, T("d1_titre"), [T("d1_desc"), "", T("x_d1_note")])

    ws = wb.create_sheet(T("x_resultats")[:31])
    cols = [T("x_section_thematique"), T("x_question"), T("x_type_reponse"),
            T("x_modalite")]
    for g in GROUPES:
        cols += [f"{g} (n)", f"{g} (%)"]
    _entete(ws, 1, cols, [28, 52, 20, 42] + [11] * (2 * len(GROUPES)))

    r = 2
    for theme in cache["themes"]:
        for label, gn in theme["rows"]:
            ws.cell(row=r, column=1, value=theme["category"]).font = NORMAL
            ws.cell(row=r, column=2, value=theme["question"]).font = NORMAL
            ws.cell(row=r, column=3, value=theme.get("note", "")).font = NORMAL
            ws.cell(row=r, column=4, value=label).font = NORMAL
            j = 5
            for g in GROUPES:
                n = gn.get(g, 0)
                b = base_n.get(g, 0)
                ws.cell(row=r, column=j, value=n).font = NORMAL
                c = ws.cell(row=r, column=j + 1,
                            value=round(n / b * 100, 1) if b else None)
                c.font = NORMAL
                c.number_format = "0.0"
                j += 2
            r += 1

    ws2 = wb.create_sheet(T("x_effectifs")[:31])
    _entete(ws2, 1, [T("x_groupe"), T("x_base_n")], [28, 14])
    for k, g in enumerate(GROUPES, 2):
        ws2.cell(row=k, column=1, value=g).font = NORMAL
        ws2.cell(row=k, column=2, value=base_n.get(g, 0)).font = NORMAL
    return _octets(wb)


# ----------------------------------------------------------------------
# 2 · indicateurs de résilience
# ----------------------------------------------------------------------
def _nom(r):
    if i18n.get_lang() == "fr" and r.get("indicateur_fr"):
        return r["indicateur_fr"]
    return r["indicateur"]


@st.cache_data(show_spinner=False)
def _fichier_indicateurs(lang):
    res = _lire_json("resultats.json")
    wb = _nouveau()
    _lisezmoi(wb, T("d2_titre"), [T("d2_desc"), "", T("x_d2_note")])

    ws = wb.create_sheet(T("x_indicateurs")[:31])
    cols = [T("x_ligne"), T("x_dimension"), T("x_indicateur"), T("x_ponderation"),
            T("x_sens"), T("x_metrique"), T("x_echelle")]
    for g in GROUPES:
        cols += [f"{g} — %", f"{g} — score"]
    _entete(ws, 1, cols,
            [8, 34, 46, 12, 20, 60, 60] + [11] * (2 * len(GROUPES)))

    r = 2
    for ind in res:
        ws.cell(row=r, column=1, value=ind["ligne"]).font = NORMAL
        ws.cell(row=r, column=2, value=ind["dimension"]).font = NORMAL
        ws.cell(row=r, column=3, value=_nom(ind)).font = NORMAL
        ws.cell(row=r, column=4, value=ind.get("ponderation")).font = NORMAL
        ws.cell(row=r, column=5, value=ind.get("sens", "")).font = NORMAL
        ws.cell(row=r, column=6, value=ind.get("metrique", "")).font = NORMAL
        ws.cell(row=r, column=7, value=ind.get("echelle", "")).font = NORMAL
        j = 8
        for g in GROUPES:
            c = ws.cell(row=r, column=j, value=(ind.get("valeurs") or {}).get(g))
            c.font = NORMAL
            c.number_format = "0.0"
            ws.cell(row=r, column=j + 1,
                    value=(ind.get("scores") or {}).get(g)).font = NORMAL
            j += 2
        r += 1
    return _octets(wb)


# ----------------------------------------------------------------------
# 3 · ventilation section × sous-population
# ----------------------------------------------------------------------
@st.cache_data(show_spinner=False)
def _fichier_ventilation(lang):
    res = _lire_json("resultats.json")
    vent = _lire_json("ventilation.json")
    par_ligne = {r["ligne"]: r for r in res}

    wb = _nouveau()
    _lisezmoi(wb, T("d3_titre"), [T("d3_desc"), "", T("x_d3_note")])

    ws = wb.create_sheet(T("x_ventilation")[:31])
    cols = [T("x_section"), T("x_ligne"), T("x_dimension"), T("x_indicateur")]
    for p in SOUS_POP:
        cols += [f"{p} — %", f"{p} — score", f"{p} — n"]
    _entete(ws, 1, cols, [18, 8, 34, 46] + [10] * (3 * len(SOUS_POP)))

    r = 2
    for sec, blocs in vent["sections"].items():
        for cle, bloc in blocs.items():
            ind = par_ligne.get(int(cle), {})
            ws.cell(row=r, column=1, value=sec).font = NORMAL
            ws.cell(row=r, column=2, value=int(cle)).font = NORMAL
            ws.cell(row=r, column=3, value=ind.get("dimension", "")).font = NORMAL
            ws.cell(row=r, column=4,
                    value=_nom(ind) if ind else "").font = NORMAL
            j = 5
            for p in SOUS_POP:
                c = ws.cell(row=r, column=j, value=bloc["valeurs"].get(p))
                c.font = NORMAL
                c.number_format = "0.0"
                ws.cell(row=r, column=j + 1,
                        value=bloc["scores"].get(p)).font = NORMAL
                ws.cell(row=r, column=j + 2,
                        value=bloc.get("n", {}).get(p)).font = NORMAL
                j += 3
            r += 1
    return _octets(wb)


# ----------------------------------------------------------------------
# 4 · scores composites pondérés
# ----------------------------------------------------------------------
def _pondere(paires):
    """paires : [(poids, score|None), ...] -> moyenne pondérée arrondie."""
    num = den = 0.0
    for p, s in paires:
        if s is None:
            continue
        num += p * s
        den += p
    return round(num / den, 2) if den else None


@st.cache_data(show_spinner=False)
def _fichier_composite(lang):
    res = _lire_json("resultats.json")
    vent = _lire_json("ventilation.json")
    scorables = [r for r in res
                 if r["calculable"] != "non" and not r["bareme_absent"]]
    poids = {r["ligne"]: (r["ponderation"] or 0.0) for r in res}
    par_dim = {d: [r["ligne"] for r in scorables if r["dimension"] == d]
               for d in DIM_ORDRE}
    lignes_tout = [r["ligne"] for r in scorables]

    wb = _nouveau()
    _lisezmoi(wb, T("d4_titre"), [T("d4_desc"), "", T("x_d4_note")])

    # -- feuille A : section × sous-population -----------------------------
    ws = wb.create_sheet(T("x_par_section")[:31])
    cols = [T("x_section"), T("x_sous_pop"), T("x_score_final")] + \
        [f"{i}. {d.split('.', 1)[1].strip().title()}"
         for i, d in enumerate(DIM_ORDRE, 1)]
    _entete(ws, 1, cols, [18, 22, 14] + [17] * len(DIM_ORDRE))

    r = 2
    for sec, blocs in vent["sections"].items():
        for p in SOUS_POP:
            ws.cell(row=r, column=1, value=sec).font = NORMAL
            ws.cell(row=r, column=2, value=p).font = NORMAL

            def paires(lignes):
                out = []
                for lg in lignes:
                    b = blocs.get(str(lg))
                    if b:
                        out.append((poids[lg], b["scores"].get(p)))
                return out

            c = ws.cell(row=r, column=3, value=_pondere(paires(lignes_tout)))
            c.font = GRAS
            c.number_format = "0.00"
            for k, d in enumerate(DIM_ORDRE):
                c = ws.cell(row=r, column=4 + k,
                            value=_pondere(paires(par_dim[d])))
                c.font = NORMAL
                c.number_format = "0.00"
            r += 1

    # -- feuille B : national, tous les groupes de référence ----------------
    ws2 = wb.create_sheet(T("x_national")[:31])
    _entete(ws2, 1, cols[1:], [22, 14] + [17] * len(DIM_ORDRE))
    par_ligne = {r_["ligne"]: r_ for r_ in scorables}
    r = 2
    for g in GROUPES:
        ws2.cell(row=r, column=1, value=g).font = NORMAL

        def paires_g(lignes, groupe=g):
            return [(poids[lg], (par_ligne[lg].get("scores") or {}).get(groupe))
                    for lg in lignes]

        c = ws2.cell(row=r, column=2, value=_pondere(paires_g(lignes_tout)))
        c.font = GRAS
        c.number_format = "0.00"
        for k, d in enumerate(DIM_ORDRE):
            c = ws2.cell(row=r, column=3 + k, value=_pondere(paires_g(par_dim[d])))
            c.font = NORMAL
            c.number_format = "0.00"
        r += 1
    return _octets(wb)


# ----------------------------------------------------------------------
# 5 · base individuelle anonymisée  /  6 · dictionnaire du questionnaire
# ----------------------------------------------------------------------
@st.cache_data(show_spinner=False)
def _fichier_brut():
    with open(_chemin("donnees_anonymisees.csv"), "rb") as f:
        return f.read()


@st.cache_data(show_spinner=False)
def _fichier_dictionnaire(lang):
    index = _lire_json("questions_index.json")
    with open(_chemin("cache_national.pkl"), "rb") as f:
        cache = pickle.load(f)
    wb = _nouveau()
    _lisezmoi(wb, T("d6_titre"), [T("d6_desc"), "", T("x_d6_note")])

    ws = wb.create_sheet(T("x_dictionnaire")[:31])
    _entete(ws, 1, [T("x_ordre"), T("x_section_thematique"), T("x_question"),
                    T("x_type_reponse"), T("x_nb_modalites")],
            [9, 30, 62, 26, 14])
    for r, q in enumerate(index, 2):
        theme = cache["themes"][q["i"]]
        ws.cell(row=r, column=1, value=q["i"] + 1).font = NORMAL
        ws.cell(row=r, column=2, value=q["category"]).font = NORMAL
        ws.cell(row=r, column=3, value=q["question"]).font = NORMAL
        ws.cell(row=r, column=4, value=theme.get("note", "")).font = NORMAL
        ws.cell(row=r, column=5, value=len(theme["rows"])).font = NORMAL
    return _octets(wb)


# ----------------------------------------------------------------------
# 7 · organisations communautaires de base
# ----------------------------------------------------------------------
@st.cache_data(show_spinner=False)
def _fichier_ocb(lang):
    ocb = _lire_json("ocb.json")
    wb = _nouveau()
    _lisezmoi(wb, T("d7_titre"), [T("d7_desc"), "", T("x_d7_note")])

    fr = i18n.get_lang() == "fr"
    OUI, NON = T("oui"), T("non")

    def trois(v):
        return "" if v is None else (OUI if v else NON)

    ws = wb.create_sheet(T("x_ocb_fiches")[:31])
    _entete(ws, 1, [T("o_col_nom"), T("o_f_localite"), T("x_section"),
                    T("o_col_partenariat"), T("o_col_duree"), T("o_col_note"),
                    T("o_col_soutien"), T("o_col_autorites"), T("o_col_ong_int"),
                    T("o_col_femme"), T("o_col_jeune"), T("o_f_projets"),
                    T("o_f_facteurs")],
            [56, 18, 18, 14, 14, 12, 16, 20, 20, 20, 22, 34, 34])
    for r, f in enumerate(ocb["fiches"], 2):
        for j, v in enumerate([
                f["nom"], f.get("localite") or "", f["section"],
                trois(f["partenariat"]), f["duree"] or "",
                f["note_partenariat"], trois(f["soutien"]), trois(f["autorites"]),
                trois(f["ong_int"]), trois(f["femme_direction"]),
                trois(f["jeune_direction"]), f.get("projets") or "",
                f.get("facteurs") or ""], 1):
            ws.cell(row=r, column=j, value=v).font = NORMAL

    ws2 = wb.create_sheet(T("x_ocb_indic")[:31])
    cols = [T("x_ligne"), T("x_dimension"), T("x_indicateur"), T("x_metrique")]
    for s in SECTIONS:
        cols += [f"{s} — {T('d_contenu')}", f"{s} — score", f"{s} — n"]
    cols += ["Total", "Total — score", "Total — n"]
    _entete(ws2, 1, cols, [8, 34, 46, 60] + [11] * (3 * len(SECTIONS) + 3))
    for r, ind in enumerate(ocb["indicateurs"], 2):
        ws2.cell(row=r, column=1, value=ind["ligne"]).font = NORMAL
        ws2.cell(row=r, column=2, value=ind["dimension"]).font = NORMAL
        ws2.cell(row=r, column=3,
                 value=(ind.get("indicateur_fr") if fr else ind["indicateur"])
                 ).font = NORMAL
        ws2.cell(row=r, column=4,
                 value=(ind.get("metrique_fr") if fr else ind.get("metrique"))
                 ).font = NORMAL
        j = 5
        for s in SECTIONS + ["Total"]:
            c = ws2.cell(row=r, column=j, value=ind["valeurs"].get(s))
            c.font = NORMAL; c.number_format = "0.0"
            ws2.cell(row=r, column=j + 1, value=ind["scores"].get(s)).font = NORMAL
            ws2.cell(row=r, column=j + 2, value=ind["n"].get(s)).font = NORMAL
            j += 3
    return _octets(wb)


# ----------------------------------------------------------------------
XLSX = ("application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet")


def _bloc(cle_titre, cle_desc, teinte, nom_fichier, mime, format_txt,
          fabrique, poids=None):
    with st.container(border=True):
        st.markdown(f'<div class="titre-bloc {teinte}">{T(cle_titre)}</div>',
                    unsafe_allow_html=True)
        st.markdown(
            f'<p style="font-size:14.5px;line-height:1.6;color:#3c4761;'
            f'margin:4px 0 10px">{T(cle_desc)}</p>', unsafe_allow_html=True)
        # Un fichier source absent du dépôt ne doit pas emporter tout l'onglet :
        # on signale le jeu manquant et les cinq autres restent téléchargeables.
        try:
            data = fabrique()
        except Exception as err:
            st.markdown(
                f'<p style="font-size:13px;color:#a8690a;background:#fdf3e3;'
                f'border:1px solid #f0dcb8;border-radius:10px;padding:9px 13px;'
                f'margin:0">{T("d_indispo", f=nom_fichier)}<br>'
                f'<span style="font-size:11.5px;color:#8a6a3a">'
                f'{type(err).__name__}</span></p>', unsafe_allow_html=True)
            return

        col_a, col_b = st.columns([1, 2])
        with col_a:
            st.download_button(
                f"{T('d_bouton')} · {format_txt}", data=data,
                file_name=nom_fichier, mime=mime,
                key=f"dl_{nom_fichier}_{i18n.get_lang()}",
                use_container_width=True)
        with col_b:
            st.markdown(
                f'<p style="font-size:12px;color:#6b7590;margin:8px 0 0">'
                f'{nom_fichier} · {len(data) / 1024:.0f} Ko</p>',
                unsafe_allow_html=True)


def render():
    lang = i18n.get_lang()

    # PAS DE TITRE DE PAGE : la colonne de menu marque déjà la rubrique. Le
    # sous-titre reste, lui : il dit ce que les sept fichiers ont en commun,
    # ce que « Données » ne dit pas.
    st.markdown(
        '<p style="font-size:11.5px;color:#6b7590;letter-spacing:.06em;'
        'text-transform:uppercase;margin:2px 0 6px;font-weight:600">'
        + T("d_sous_titre") + "</p>", unsafe_allow_html=True)

    st.markdown(
        '<div style="background:#fff;border:1px solid #e3eaf3;border-left:5px '
        'solid #1a6bb0;border-radius:14px;padding:13px 17px;font-size:14.5px;'
        'color:#3c4761;box-shadow:0 1px 2px rgba(16,23,40,.05),'
        '0 8px 20px rgba(16,23,40,.06);margin:10px 0 6px">'
        + T("d_intro") + "</div>", unsafe_allow_html=True)
    st.markdown(
        '<div style="background:#fdf7ec;border:1px solid #f0dcb8;border-left:5px '
        'solid #d99b28;border-radius:14px;padding:13px 17px;font-size:14.5px;'
        'color:#5b4a2b;margin:0 0 14px">' + T("d_avert") + "</div>",
        unsafe_allow_html=True)

    _bloc("d1_titre", "d1_desc", "", "01_resultats_descriptifs.xlsx", XLSX,
          "Excel", lambda: _fichier_descriptif(lang))
    _bloc("d2_titre", "d2_desc", "vert", "02_indicateurs_resilience.xlsx", XLSX,
          "Excel", lambda: _fichier_indicateurs(lang))
    _bloc("d3_titre", "d3_desc", "ambre", "03_ventilation_section_souspop.xlsx",
          XLSX, "Excel", lambda: _fichier_ventilation(lang))
    _bloc("d4_titre", "d4_desc", "", "04_scores_composites.xlsx", XLSX,
          "Excel", lambda: _fichier_composite(lang))
    _bloc("d5_titre", "d5_desc", "vert", "05_base_individuelle_anonymisee.csv",
          "text/csv", "CSV", _fichier_brut)
    _bloc("d6_titre", "d6_desc", "ambre", "06_dictionnaire_questionnaire.xlsx",
          XLSX, "Excel", lambda: _fichier_dictionnaire(lang))
    _bloc("d7_titre", "d7_desc", "", "07_organisations_communautaires.xlsx",
          XLSX, "Excel", lambda: _fichier_ocb(lang))

    st.caption(T("credit"))
